"""
Excel Data Analysis Portfolio
Author: Bienvenu Mwenyemali
Skills Demonstrated: Excel Formulas, VLOOKUP, Pivot Tables, Charts, Dashboard
"""

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter
from datetime import datetime

print("=" * 60)
print("CREATING EXCEL DATA ANALYSIS PORTFOLIO")
print("=" * 60)

# Fetch sample data from API
print("\n1. Fetching data from API...")
response = requests.get("https://smartcoinsapp.com/api/coins", timeout=30)
api_data = response.json().get('data', [])[:30]  # Use only 30 coins to keep file light
print(f"   Retrieved {len(api_data)} coins (sample)")

# Transform to DataFrame
records = []
for coin in api_data:
    inv = coin.get('investmentScores', {})
    records.append({
        'Coin': coin.get('name'),
        'Symbol': coin.get('symbol'),
        'Price': round(coin.get('price', 0), 6),
        'MarketCap': round(coin.get('marketCap', 0), 2),
        'Volume24h': round(coin.get('volume24h', 0), 2),
        'Change24h': round(coin.get('percentChange24h', 0), 2),
        'Change7d': round(coin.get('percentChange7d', 0), 2),
        'Change30d': round(coin.get('percentChange30d', 0), 2),
        'Type': coin.get('coinType', 'Unknown'),
        'Category': coin.get('category', 'Unknown'),
        'Signal': coin.get('primarySignal', 'NEUTRAL'),
        'SignalStrength': coin.get('signalStrength', 0),
        'OverallScore': round(coin.get('overallScore', 0), 2),
        'VolatilityRisk': round(coin.get('volatilityRisk', 0), 2),
        'LiquidityRisk': round(coin.get('liquidityRisk', 0), 2),
        'MomentumScore': round(inv.get('momentum', 0), 2),
        'RiskScore': round(inv.get('risk', 0), 2),
    })

df = pd.DataFrame(records)
print(f"   Created DataFrame: {len(df)} rows")

# Create Workbook
wb = Workbook()
print("\n2. Creating Excel sheets...")

# ============================================================================
# SHEET 1: RAW DATA
# ============================================================================
ws_data = wb.active
ws_data.title = "RawData"

# Add title
ws_data['A1'] = "SMARTCOINS RAW DATA"
ws_data['A1'].font = Font(bold=True, size=14)
ws_data.merge_cells('A1:Q1')

# Add headers and data starting from row 3
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 3:  # Header row
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

# Auto-width columns
for col in range(1, len(df.columns) + 1):
    ws_data.column_dimensions[get_column_letter(col)].width = 15

print("   - RawData sheet created")

# ============================================================================
# SHEET 2: LOOKUP TABLE (Reference for VLOOKUP)
# ============================================================================
ws_lookup = wb.create_sheet("LookupTable")

# Signal Reference Table
ws_lookup['A1'] = "SIGNAL REFERENCE TABLE"
ws_lookup['A1'].font = Font(bold=True, size=12)
ws_lookup.merge_cells('A1:C1')

headers = ['Signal', 'Interpretation', 'ActionCode']
signals = [
    ['STRONG_BUY', 'Very bullish - High confidence buy', 5],
    ['BUY', 'Bullish - Consider buying', 4],
    ['NEUTRAL', 'Hold position - Wait for signals', 3],
    ['SELL', 'Bearish - Consider selling', 2],
    ['STRONG_SELL', 'Very bearish - High confidence sell', 1]
]

for c, h in enumerate(headers, 1):
    cell = ws_lookup.cell(row=3, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

for r, row in enumerate(signals, 4):
    for c, val in enumerate(row, 1):
        ws_lookup.cell(row=r, column=c, value=val)

# Risk Level Table
ws_lookup['E1'] = "RISK LEVEL TABLE"
ws_lookup['E1'].font = Font(bold=True, size=12)
ws_lookup.merge_cells('E1:G1')

risk_headers = ['MinRisk', 'MaxRisk', 'RiskLevel']
risk_levels = [
    [0, 0.5, 'Low Risk'],
    [0.5, 2, 'Medium Risk'],
    [2, 100, 'High Risk']
]

for c, h in enumerate(risk_headers, 5):
    cell = ws_lookup.cell(row=3, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")

for r, row in enumerate(risk_levels, 4):
    for c, val in enumerate(row, 5):
        ws_lookup.cell(row=r, column=c, value=val)

# Category Reference
ws_lookup['I1'] = "CATEGORY REFERENCE"
ws_lookup['I1'].font = Font(bold=True, size=12)

cat_headers = ['Category', 'Description']
categories = [
    ['Large', 'High market cap coins'],
    ['Medium', 'Mid-sized market cap'],
    ['Small', 'Low market cap - Higher risk'],
    ['Micro', 'Very small - Speculative']
]

for c, h in enumerate(cat_headers, 9):
    cell = ws_lookup.cell(row=3, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

for r, row in enumerate(categories, 4):
    for c, val in enumerate(row, 9):
        ws_lookup.cell(row=r, column=c, value=val)

for col in ['A', 'B', 'C', 'E', 'F', 'G', 'I', 'J']:
    ws_lookup.column_dimensions[col].width = 18

print("   - LookupTable sheet created")

# ============================================================================
# SHEET 3: FORMULAS DEMONSTRATION
# ============================================================================
ws_formulas = wb.create_sheet("FormulasDemo")

ws_formulas['A1'] = "EXCEL FORMULAS DEMONSTRATION"
ws_formulas['A1'].font = Font(bold=True, size=14)
ws_formulas.merge_cells('A1:F1')

# Copy subset of data for formula demo
ws_formulas['A3'] = "Symbol"
ws_formulas['B3'] = "Price"
ws_formulas['C3'] = "OverallScore"
ws_formulas['D3'] = "Signal"
ws_formulas['E3'] = "VolatilityRisk"

for c in range(1, 6):
    ws_formulas.cell(row=3, column=c).font = Font(bold=True, color="FFFFFF")
    ws_formulas.cell(row=3, column=c).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Add sample data (first 15 rows)
for i, row in df.head(15).iterrows():
    ws_formulas.cell(row=i+4, column=1, value=row['Symbol'])
    ws_formulas.cell(row=i+4, column=2, value=row['Price'])
    ws_formulas.cell(row=i+4, column=3, value=row['OverallScore'])
    ws_formulas.cell(row=i+4, column=4, value=row['Signal'])
    ws_formulas.cell(row=i+4, column=5, value=row['VolatilityRisk'])

# Formula Examples Section
ws_formulas['G3'] = "FORMULA EXAMPLES"
ws_formulas['G3'].font = Font(bold=True, size=12, color="FFFFFF")
ws_formulas['G3'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws_formulas.merge_cells('G3:I3')

formulas = [
    ('G5', 'SUM (Total Score):', 'H5', '=SUM(C4:C18)'),
    ('G6', 'AVERAGE (Avg Score):', 'H6', '=AVERAGE(C4:C18)'),
    ('G7', 'MAX (Highest Score):', 'H7', '=MAX(C4:C18)'),
    ('G8', 'MIN (Lowest Score):', 'H8', '=MIN(C4:C18)'),
    ('G9', 'COUNT (Total Coins):', 'H9', '=COUNT(C4:C18)'),
    ('G10', 'COUNTIF (BUY Signals):', 'H10', '=COUNTIF(D4:D18,"*BUY*")'),
    ('G11', 'SUMIF (Score if BUY):', 'H11', '=SUMIF(D4:D18,"*BUY*",C4:C18)'),
    ('G12', 'AVERAGEIF (Avg if BUY):', 'H12', '=AVERAGEIF(D4:D18,"*BUY*",C4:C18)'),
]

for label_cell, label, formula_cell, formula in formulas:
    ws_formulas[label_cell] = label
    ws_formulas[label_cell].font = Font(bold=True)
    ws_formulas[formula_cell] = formula

# VLOOKUP Example
ws_formulas['G14'] = "VLOOKUP EXAMPLE"
ws_formulas['G14'].font = Font(bold=True, size=11, color="FFFFFF")
ws_formulas['G14'].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
ws_formulas.merge_cells('G14:I14')

ws_formulas['G15'] = "Lookup Symbol:"
ws_formulas['H15'] = df.iloc[0]['Symbol']  # First symbol for demo
ws_formulas['G16'] = "Found Signal:"
ws_formulas['H16'] = f'=VLOOKUP(H15,A4:E18,4,FALSE)'
ws_formulas['G17'] = "Found Score:"
ws_formulas['H17'] = f'=VLOOKUP(H15,A4:E18,3,FALSE)'

# IF Statement Example
ws_formulas['G19'] = "IF STATEMENT EXAMPLE"
ws_formulas['G19'].font = Font(bold=True, size=11, color="FFFFFF")
ws_formulas['G19'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
ws_formulas.merge_cells('G19:I19')

ws_formulas['G20'] = "Risk Classification:"
ws_formulas['H20'] = '=IF(E4<0.5,"Low",IF(E4<2,"Medium","High"))'

# INDEX MATCH Example
ws_formulas['G22'] = "INDEX MATCH (Better VLOOKUP)"
ws_formulas['G22'].font = Font(bold=True, size=11, color="FFFFFF")
ws_formulas['G22'].fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
ws_formulas.merge_cells('G22:I22')

ws_formulas['G23'] = "Find Price by Symbol:"
ws_formulas['H23'] = '=INDEX(B4:B18,MATCH(H15,A4:A18,0))'

for col in ['A', 'B', 'C', 'D', 'E', 'G', 'H', 'I']:
    ws_formulas.column_dimensions[col].width = 18

print("   - FormulasDemo sheet created")

# ============================================================================
# SHEET 4: PIVOT TABLE STYLE SUMMARY
# ============================================================================
ws_pivot = wb.create_sheet("PivotSummary")

ws_pivot['A1'] = "PIVOT TABLE STYLE SUMMARIES"
ws_pivot['A1'].font = Font(bold=True, size=14)
ws_pivot.merge_cells('A1:E1')

# Summary by Signal
ws_pivot['A3'] = "SUMMARY BY SIGNAL"
ws_pivot['A3'].font = Font(bold=True, size=12)

signal_summary = df.groupby('Signal').agg({
    'Coin': 'count',
    'OverallScore': 'mean',
    'VolatilityRisk': 'mean',
    'Price': 'mean'
}).round(2).reset_index()
signal_summary.columns = ['Signal', 'Count', 'AvgScore', 'AvgRisk', 'AvgPrice']

for c, col in enumerate(signal_summary.columns, 1):
    cell = ws_pivot.cell(row=4, column=c, value=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for r, row in enumerate(signal_summary.values, 5):
    for c, val in enumerate(row, 1):
        ws_pivot.cell(row=r, column=c, value=val)

# Summary by Type
ws_pivot['A12'] = "SUMMARY BY COIN TYPE"
ws_pivot['A12'].font = Font(bold=True, size=12)

type_summary = df.groupby('Type').agg({
    'Coin': 'count',
    'OverallScore': 'mean',
    'MarketCap': 'sum'
}).round(2).reset_index()
type_summary.columns = ['Type', 'Count', 'AvgScore', 'TotalMarketCap']

for c, col in enumerate(type_summary.columns, 1):
    cell = ws_pivot.cell(row=13, column=c, value=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

for r, row in enumerate(type_summary.values, 14):
    for c, val in enumerate(row, 1):
        ws_pivot.cell(row=r, column=c, value=val)

# Summary by Category
ws_pivot['A22'] = "SUMMARY BY CATEGORY"
ws_pivot['A22'].font = Font(bold=True, size=12)

cat_summary = df.groupby('Category').agg({
    'Coin': 'count',
    'OverallScore': 'mean',
    'VolatilityRisk': 'mean'
}).round(2).reset_index()
cat_summary.columns = ['Category', 'Count', 'AvgScore', 'AvgRisk']

for c, col in enumerate(cat_summary.columns, 1):
    cell = ws_pivot.cell(row=23, column=c, value=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")

for r, row in enumerate(cat_summary.values, 24):
    for c, val in enumerate(row, 1):
        ws_pivot.cell(row=r, column=c, value=val)

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_pivot.column_dimensions[col].width = 18

print("   - PivotSummary sheet created")

# ============================================================================
# SHEET 5: DASHBOARD
# ============================================================================
ws_dash = wb.create_sheet("Dashboard")

# Title
ws_dash['A1'] = "SMARTCOINS ANALYSIS DASHBOARD"
ws_dash['A1'].font = Font(bold=True, size=16)
ws_dash.merge_cells('A1:L1')

ws_dash['A2'] = f"Analysis Date: {datetime.now().strftime('%Y-%m-%d')} | Author: Bienvenu Mwenyemali"
ws_dash['A2'].font = Font(italic=True, size=10)

# KPI Section
kpis = [
    ('A4', 'Total Coins', len(df)),
    ('C4', 'Avg Score', round(df['OverallScore'].mean(), 2)),
    ('E4', 'BUY Signals', len(df[df['Signal'].str.contains('BUY', na=False)])),
    ('G4', 'Low Risk', len(df[df['VolatilityRisk'] < 0.5])),
    ('I4', 'Total Market Cap', f"${df['MarketCap'].sum():,.0f}"),
]

for cell, label, value in kpis:
    ws_dash[cell] = label
    ws_dash[cell].font = Font(bold=True, size=10)
    ws_dash[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_dash[cell].font = Font(bold=True, color="FFFFFF")
    
    value_cell = cell[0] + str(int(cell[1]) + 1)
    ws_dash[value_cell] = value
    ws_dash[value_cell].font = Font(bold=True, size=14)
    ws_dash[value_cell].alignment = Alignment(horizontal="center")

# Top 10 Coins Table
ws_dash['A8'] = "TOP 10 COINS BY SCORE"
ws_dash['A8'].font = Font(bold=True, size=12)

top10 = df.nlargest(10, 'OverallScore')[['Coin', 'Symbol', 'OverallScore', 'Signal']].reset_index(drop=True)

for c, col in enumerate(top10.columns, 1):
    cell = ws_dash.cell(row=9, column=c, value=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

for r, row in enumerate(top10.values, 10):
    for c, val in enumerate(row, 1):
        ws_dash.cell(row=r, column=c, value=val)

# Signal Distribution for Chart
ws_dash['F8'] = "SIGNAL DISTRIBUTION"
ws_dash['F8'].font = Font(bold=True, size=12)

signal_dist = df['Signal'].value_counts().reset_index()
signal_dist.columns = ['Signal', 'Count']

for c, col in enumerate(signal_dist.columns, 6):
    cell = ws_dash.cell(row=9, column=c, value=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")

for r, row in enumerate(signal_dist.values, 10):
    for c, val in enumerate(row, 6):
        ws_dash.cell(row=r, column=c, value=val)

# Create Bar Chart for Top 10
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Top 10 Coins by Score"
chart1.y_axis.title = "Score"

data = Reference(ws_dash, min_col=3, min_row=9, max_row=19)
cats = Reference(ws_dash, min_col=2, min_row=10, max_row=19)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
chart1.width = 12
chart1.height = 8

ws_dash.add_chart(chart1, "A22")

# Create Pie Chart for Signal Distribution
chart2 = PieChart()
chart2.title = "Signal Distribution"

data2 = Reference(ws_dash, min_col=7, min_row=9, max_row=9+len(signal_dist))
cats2 = Reference(ws_dash, min_col=6, min_row=10, max_row=9+len(signal_dist))
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width = 10
chart2.height = 8

ws_dash.add_chart(chart2, "H22")

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws_dash.column_dimensions[col].width = 14

print("   - Dashboard sheet created")

# ============================================================================
# SHEET 6: CONDITIONAL FORMATTING DEMO
# ============================================================================
ws_cond = wb.create_sheet("ConditionalFormat")

ws_cond['A1'] = "CONDITIONAL FORMATTING DEMONSTRATION"
ws_cond['A1'].font = Font(bold=True, size=14)
ws_cond.merge_cells('A1:E1')

# Add data
headers = ['Coin', 'Score', 'Risk', 'Change24h', 'Signal']
for c, h in enumerate(headers, 1):
    cell = ws_cond.cell(row=3, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for i, row in df.head(20).iterrows():
    ws_cond.cell(row=i+4, column=1, value=row['Coin'])
    ws_cond.cell(row=i+4, column=2, value=row['OverallScore'])
    ws_cond.cell(row=i+4, column=3, value=row['VolatilityRisk'])
    ws_cond.cell(row=i+4, column=4, value=row['Change24h'])
    ws_cond.cell(row=i+4, column=5, value=row['Signal'])

# Add Color Scale to Score column (B)
ws_cond.conditional_formatting.add('B4:B23',
    ColorScaleRule(start_type='min', start_color='F8696B',
                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                   end_type='max', end_color='63BE7B'))

# Add Color Scale to Risk column (C) - reversed
ws_cond.conditional_formatting.add('C4:C23',
    ColorScaleRule(start_type='min', start_color='63BE7B',
                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                   end_type='max', end_color='F8696B'))

# Explanation
ws_cond['G3'] = "FORMATTING APPLIED:"
ws_cond['G3'].font = Font(bold=True)
ws_cond['G4'] = "- Score: Green (high) to Red (low)"
ws_cond['G5'] = "- Risk: Green (low) to Red (high)"
ws_cond['G6'] = "- Change24h: Shows positive/negative"
ws_cond['G8'] = "This helps quickly identify:"
ws_cond['G9'] = "- Best performing coins (green score)"
ws_cond['G10'] = "- Lowest risk coins (green risk)"
ws_cond['G11'] = "- Price movements at a glance"

for col in ['A', 'B', 'C', 'D', 'E', 'G']:
    ws_cond.column_dimensions[col].width = 16

print("   - ConditionalFormat sheet created")

# ============================================================================
# SAVE WORKBOOK
# ============================================================================
output_path = "output/SmartCoins_Excel_Analysis.xlsx"
wb.save(output_path)

print(f"\n3. Excel file saved: {output_path}")

print("\n" + "=" * 60)
print("EXCEL FILE CREATED SUCCESSFULLY!")
print("=" * 60)
print("\nSheets included:")
print("  1. RawData - Original data from API")
print("  2. LookupTable - Reference tables for VLOOKUP")
print("  3. FormulasDemo - SUM, AVERAGE, VLOOKUP, IF, INDEX/MATCH")
print("  4. PivotSummary - Pivot table style summaries")
print("  5. Dashboard - KPIs, charts, top performers")
print("  6. ConditionalFormat - Color scales and formatting")
print("\nSkills Demonstrated:")
print("  - Data import and organization")
print("  - VLOOKUP and INDEX/MATCH")
print("  - Aggregate functions (SUM, AVERAGE, COUNT)")
print("  - COUNTIF, SUMIF, AVERAGEIF")
print("  - IF statements and nested IFs")
print("  - Pivot table summaries")
print("  - Dashboard with KPIs")
print("  - Bar charts and Pie charts")
print("  - Conditional formatting (color scales)")
print("=" * 60)
